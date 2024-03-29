# Steps
"""
- request for input
- search for occurrences in source file
- copy entire row of occurrence to new file
- when done, ask for input again
- repeat process on new sheet. 
"""
import openpyxl
import sys
from openpyxl.styles import cell_style, Font


def style(file):
    """ applies the normal styling to cells in the provided file (path)"""
    source = openpyxl.load_workbook(file)
    sheet = source.active
    
    # applying a fixed width to each prescribed width to every column
    sheet.row_dimensions[1].font = Font(bold=True)
    

def match(pattern):
    population = ["Aiico", "FPML", "NLPC", "Access", "African Alliance", "APT", "ARM", "GREAT", "Assurance Annunity", "AXA", "Chevron", "CornerStone", "Coronation", "CPL", "Crusader", "Fidelity", "Great Nigeria", "IPML", "Leadway", "Niger Insurance", "Norrenberger", "NPF", "NUPEMCO", "OAK", "PAL", "PPL", "Radix", "SNCPFA", "Trust", "Veritas"]
    
    for i in range(0, len(population)):
        if pattern in population[i]:
            return population[i]
        elif pattern.capitalize() in population[i]:
            return population[i]
        elif pattern.upper() in population[i]:
            return population[i]
        elif pattern.lower() in population[i]:
            return population[i]
    
def main():
    count = 0

    # load a workbook
    source = openpyxl.load_workbook("./src/BALANCES 29.12.2023.xlsx")   # input relative path to workbook
    sheet = source.active

    connector = []
    current_list = []

    # main loop
    while True:
        while count >= 1:
            re_prompt = input("Do you want to sort for another company? (y/n) ")

            re_prompt = re_prompt.strip().lower()

            if re_prompt == "y" or re_prompt == "yes":
                break

            elif re_prompt == "n" or re_prompt == "no":
                sys.exit()

            else:
                print("Provide a valid answer! ")
                continue

        pattern = input("What do you wish to sort? ")

        # input validation
        pattern.strip()    # note: add support for fuzzy search and combinatory integration.
        pattern = match(pattern)
        
        if type(pattern) != str:
            print("We do not know what that is.\nDo try again later. ")
            break
        
        # special consideration for contingent names
        if pattern == "PPL":
            alt = "Prem"
            alt2 = "PREM"
        elif pattern == "Veritas":
            alt = "VG"
        
            
            
        new_book = openpyxl.Workbook()
        new_sheet = new_book.active

        # search for pattern in file and store them
        for rows in range(1, 10001):
            # condition for EOF: Three consecutive empty lines
            if sheet[f"C{rows}"].value is None and sheet[f"C{rows + 1}"].value is None:  # if the iterator comes on two empty lines
                if sheet[f"C{rows + 2}"].value is None:  # check if the following line is also empty.
                    break  # break out of loop

            # logic for crossing domains
            # if sheet[f"C{rows}"].value is None and sheet[f"C{rows + 2}"].value is not None:
            #     # connector.append([])
            #     # connector.append(list(sheet[f"C{rows + 1}"].value))
            #     # connector.append([])
            #     ...

            # finding and copying the heading
            if "Participant" in sheet[f"C{rows}"].value and sheet[f"C{rows}"].value is not None:
                current = list(sheet[rows])
                for i in current:
                    current_list.append(i.value)

                # looping through the current list inorder to append its values independently to the connector
                connector.append([])    # create an empty list to house the heading
                for i in range(len(current_list)):
                    connector[0].append(current_list[i])
                current_list.clear()

            # logic for grepping pattern
            if pattern != "PPL" or pattern != "Veritas":
    
                if sheet[f"C{rows}"].value is None and sheet[f"C{rows + 1}"].value is not None:
                    continue
                elif pattern in sheet[f"C{rows}"].value and sheet[f"C{rows}"].value is not None:
                    current = list(sheet[rows])
                    for i in current:
                        current_list.append(i.value)

                    # looping through the current list inorder to append its values independently to the connector
                    connector.append([])  # create an empty list to house the heading
                    for i in current_list:
                        connector[len(connector) - 1].append(i)
                    current_list.clear()
                    
            else:
                if sheet[f"C{rows}"].value is None and sheet[f"C{rows + 1}"].value is not None:
                    continue
                elif pattern in sheet[f"C{rows}"].value or alt in sheet[f"C{rows}"].value or alt2 in sheet[f"C{rows}"].value and sheet[f"C{rows}"].value is not None:
                    current = list(sheet[rows])
                    for i in current:
                        current_list.append(i.value)

                    # looping through the current list inorder to append its values independently to the connector
                    connector.append([])  # create an empty list to house the heading
                    for i in current_list:
                        connector[len(connector) - 1].append(i)
                    current_list.clear()
                

        # error checking for final step
        if len(connector) < 2:
            print("{} not found!".format(pattern))
            connector.clear()
        else:
            # writing to new file from storage
            for i in connector:
                new_sheet.append(i)  # new worksheet uses just as many columns as required.
                
            new_book.save(f"./output/{pattern}.xlsx")
            print("Successfully sorted {}".format(pattern))
            print("Path to sorted file is output/{}.xlsx".format(pattern))
            connector.clear()
            
            # formatting and error checking
            try:
                file = f"./output/{pattern}.xlsx"
                style(file)
            
            except AttributeError:
                print(f"Couldn't format {pattern}.xlsx")

        count += 1


if __name__ == "__main__":
    main()
